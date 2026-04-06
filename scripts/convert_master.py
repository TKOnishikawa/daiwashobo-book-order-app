"""xlsx マスタ → book_master.json 変換スクリプト
openBD API で30文字切れタイトルを自動補完する。
"""
import sys, io, json, time, urllib.request
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

import openpyxl

INPUT = r"c:\Users\takao\Downloads\0325在庫+直近１年実績表（正）.xlsx"
OUTPUT = r"c:\Users\takao\OneDrive\ドキュメント\GitHub\dx-projects\clients\daiwashobo\book-order-app\public\book_master.json"

PREFIX = "9784479"  # 大和書房の出版者コード (978-4-479)

TITLE_TRUNCATE_LEN = 30  # Excel側の切り捨て文字数

# 全角→半角変換テーブル
ZEN2HAN = str.maketrans(
    "０１２３４５６７８９ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ",
    "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ",
)


def normalize_size(raw: str) -> str:
    if not raw:
        return ""
    s = raw.translate(ZEN2HAN).strip()
    if "文庫" in s and "6" in s:
        return "文庫版"
    return s


def normalize_author(raw: str) -> str:
    if not raw:
        return ""
    return raw.replace(" ", "").strip()


def fetch_openbd_titles(isbns: list[str]) -> dict[str, str]:
    """openBD APIでISBN→タイトルを一括取得。1リクエスト最大1000件。"""
    result = {}
    batch_size = 1000
    for start in range(0, len(isbns), batch_size):
        batch = isbns[start : start + batch_size]
        url = f"https://api.openbd.jp/v1/get?isbn={','.join(batch)}"
        try:
            with urllib.request.urlopen(url, timeout=30) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            for item in data:
                if item:
                    summary = item.get("summary", {})
                    isbn = summary.get("isbn", "")
                    title = summary.get("title", "")
                    if isbn and title:
                        result[isbn] = title
        except Exception as e:
            print(f"  openBD batch error: {e}")
        if start + batch_size < len(isbns):
            time.sleep(1)  # rate limit courtesy
    return result


def main():
    print(f"Reading: {INPUT}")
    wb = openpyxl.load_workbook(INPUT, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Phase 1: Extract records from Excel
    records = []
    truncated_isbns = []

    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        code = str(row[0] or "").strip()
        check = str(row[1] or "").strip()
        title = str(row[2] or "").strip()
        author = normalize_author(str(row[3] or ""))
        price = str(int(float(str(row[4])))) if row[4] else ""
        size = normalize_size(str(row[41] or ""))    # 版型1 = col AR (index 41)
        pages = str(int(float(str(row[48])))) if row[48] else ""  # ページ数 = col AW (index 48)
        genre = str(row[51] or "").strip()            # 商品分類 = col AZ (index 51)

        if not code or not title:
            continue

        isbn13 = PREFIX + code.zfill(5) + check
        if len(isbn13) != 13:
            print(f"  WARN: Row {i+2} ISBN length {len(isbn13)}: {isbn13}")
            continue

        # Flag truncated titles
        is_truncated = len(title) >= TITLE_TRUNCATE_LEN

        records.append({
            "isbn13": isbn13,
            "title": title,
            "author": author,
            "price": price,
            "genre": genre,
            "size": size,
            "pages": pages,
            "_truncated": is_truncated,
        })

        if is_truncated:
            truncated_isbns.append(isbn13)

    wb.close()
    print(f"Extracted {len(records)} records, {len(truncated_isbns)} potentially truncated titles")

    # Phase 2: Fetch full titles from openBD for truncated ones
    if truncated_isbns:
        print(f"Fetching {len(truncated_isbns)} titles from openBD...")
        openbd_titles = fetch_openbd_titles(truncated_isbns)
        print(f"  Got {len(openbd_titles)} titles from openBD")

        fixed = 0
        not_found = []
        for rec in records:
            if rec["_truncated"] and rec["isbn13"] in openbd_titles:
                old_title = rec["title"]
                new_title = openbd_titles[rec["isbn13"]]
                if len(new_title) > len(old_title):
                    rec["title"] = new_title
                    fixed += 1
                    print(f"  Fixed: [{old_title}] → [{new_title}]")
            elif rec["_truncated"] and rec["isbn13"] not in openbd_titles:
                not_found.append(f"{rec['isbn13']}: {rec['title']}")

        print(f"  Fixed {fixed} titles, {len(not_found)} not found in openBD")
        if not_found:
            print("  Not found:")
            for nf in not_found[:10]:
                print(f"    {nf}")
            if len(not_found) > 10:
                print(f"    ... and {len(not_found) - 10} more")

    # Phase 3: Clean up and write
    for rec in records:
        del rec["_truncated"]

    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False)

    print(f"\nWrote {len(records)} records → {OUTPUT}")

    # Verify
    verify_count_30 = sum(1 for r in records if len(r["title"]) == TITLE_TRUNCATE_LEN)
    verify_count_over = sum(1 for r in records if len(r["title"]) > TITLE_TRUNCATE_LEN)
    print(f"Verification: titles ==30: {verify_count_30}, >30: {verify_count_over}")


if __name__ == "__main__":
    main()
